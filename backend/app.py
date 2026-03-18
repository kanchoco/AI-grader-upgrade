import os
from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
from google.cloud.sql.connector import Connector
import sqlalchemy
import pandas as pd
import uuid
import json
from ai_grader import run_ai_grading
from flask import send_file
import io
from collections import defaultdict
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter



# React build 경로
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
FRONTEND_BUILD_PATH = os.path.join(BASE_DIR, "dist")

print(BASE_DIR)
print(FRONTEND_BUILD_PATH)
print(os.path.exists(FRONTEND_BUILD_PATH))

# Flask app
app = Flask(
    __name__,
    static_folder=FRONTEND_BUILD_PATH,
    static_url_path=""
)

CORS(app)

# 환경변수 (Cloud Run)
DB_USER = os.environ["DB_USER"]
DB_PASS = os.environ["DB_PASS"]
DB_NAME = os.environ["DB_NAME"]
CONN_NAME = os.environ["CONN_NAME"] 

connector = Connector()

# Cloud SQL 연결
def get_engine():
    def getconn():
        return connector.connect(
            CONN_NAME,
            "pymysql",
            user=DB_USER,
            password=DB_PASS,
            db=DB_NAME
        )

    return sqlalchemy.create_engine(
        "mysql+pymysql://",
        creator=getconn,
        pool_pre_ping=True,
    )

# API 영역
@app.route("/")
def serve():
    return app.send_static_file("index.html")

@app.post("/upload_excel")
def upload_excel():
    if "file" not in request.files:
        return {"status": "error", "message": "No file uploaded"}, 400

    file = request.files["file"]

    project_name = request.form.get("projectName")
    name_column = request.form.get("nameColumn")
    answer_column = request.form.get("answerColumn")
    criteria_raw = request.form.get("criteria")
    rubric = request.form.get("rubric", "")

    if not project_name:
        return {"status": "error", "message": "프로젝트명 누락"}, 400

    if not name_column or not answer_column:
        return {"status": "error", "message": "이름/답변 열 정보 누락"}, 400

    if not criteria_raw:
        return {"status": "error", "message": "criteria 누락"}, 400

    try:
        criteria = json.loads(criteria_raw)
    except:
        return {"status": "error", "message": "criteria JSON 파싱 실패"}, 400

    try:
        df = pd.read_excel(file)

        # ---- 동적 컬럼 검사 ----
        if name_column not in df.columns:
            return {
                "status": "error",
                "message": f"엑셀에 해당 이름 열이 없습니다: {name_column}"
            }, 400

        if answer_column not in df.columns:
            return {
                "status": "error",
                "message": f"엑셀에 해당 답변 열이 없습니다: {answer_column}"
            }, 400

        engine = get_engine()

        with engine.connect() as conn:

            # ---- 프로젝트 생성 ----
            result = conn.execute(
                sqlalchemy.text("""
                    INSERT INTO projectDB
                    (project_name, criteria, prompt_text, created_at)
                    VALUES (:name, :criteria, :prompt_text, NOW())
                """),
                {
                    "name": project_name,
                    "criteria": json.dumps(criteria),
                    "prompt_text": rubric
                }
            )

            project_id = result.lastrowid

            inserted = 0

            for _, row in df.iterrows():
                conn.execute(
                    sqlalchemy.text("""
                        INSERT INTO studentDB
                        (project_id, student_name, student_answer, created_at)
                        VALUES (:pid, :name, :answer, NOW())
                    """),
                    {
                        "pid": project_id,
                        "name": str(row[name_column]).strip(),
                        "answer": str(row[answer_column]).strip()
                    }
                )
                inserted += 1

            conn.commit()

        return {
            "status": "success",
            "message": f"{inserted} students added",
            "project_id": project_id
        }

    except Exception as e:
        return {"status": "error", "message": str(e)}, 500
    

def get_project_id(conn, project_name):
    row = conn.execute(
        sqlalchemy.text("""
            SELECT project_id
            FROM projectDB
            WHERE project_name = :name
        """),
        {"name": project_name}
    ).mappings().fetchone()

    if not row:
        return None

    return row["project_id"]

def get_feedback_rows(conn, project_id):
    return conn.execute(
        sqlalchemy.text("""
            SELECT 
                s.student_name,
                f.criterion_name,
                f.rationale,
                f.key_sentence
            FROM ai_feedback_log f
            JOIN studentDB s 
                ON f.student_id = s.student_id
            WHERE f.project_id = :project_id
        """),
        {"project_id": project_id}
    ).mappings().all()

def get_score_rows(conn, project_id):
    return conn.execute(
        sqlalchemy.text("""
            SELECT 
                s.student_name,
                r.rater_name,
                sc.stage,
                sc.scores
            FROM scoreDB sc
            JOIN studentDB s 
                ON sc.student_id = s.student_id
            JOIN raterDB r 
                ON sc.rater_uid = r.rater_uid
            WHERE sc.project_id = :project_id
        """),
        {"project_id": project_id}
    ).mappings().all()

def format_text(text):
    if not text:
        return ""

    text = text.replace(". ", ".\n")

    text = text.replace("- ", "\n- ")

    return text.strip()

def sort_columns(cols):
    ordered = []

    if "student_name" in cols:
        ordered.append("student_name")
    if "rater_name" in cols:
        ordered.append("rater_name")

    criteria = set()
    for c in cols:
        if "_" in c:
            crit = c.rsplit("_", 1)[0]
            criteria.add(crit)

    for crit in sorted(criteria):
        for stage in ["human", "ai", "final"]:
            col = f"{crit}_{stage}"
            if col in cols:
                ordered.append(col)

        if f"{crit}_rationale" in cols:
            ordered.append(f"{crit}_rationale")
        if f"{crit}_evidence" in cols:
            ordered.append(f"{crit}_evidence")

    return ordered


@app.get("/export_project_excel/<project_name>")
def export_project_excel(project_name):

    engine = get_engine()

    with engine.connect() as conn:

        project_id = get_project_id(conn, project_name)

        if not project_id:
            return {"success": False, "message": "project not found"}, 404
        
        score_rows = get_score_rows(conn, project_id)

        merged = defaultdict(lambda: defaultdict(dict))

        for row in score_rows:
            key = row["student_name"]
            scores = json.loads(row["scores"])

            for criterion, value in scores.items():
                merged[key][criterion][row["stage"]] = value

            merged[key]["student_name"] = row["student_name"]
            merged[key]["rater_name"] = row["rater_name"]

        rows = []

        for key, data in merged.items():
            base = {
                "student_name": data["student_name"],
                "rater_name": data["rater_name"]
            }

            for criterion, stages in data.items():
                if criterion in ["student_name", "rater_name"]:
                    continue

                base[f"{criterion}_human"] = stages.get("human", "")
                base[f"{criterion}_ai"] = stages.get("ai", "")
                base[f"{criterion}_final"] = stages.get("final", "")

            rows.append(base)

        feedback_rows = get_feedback_rows(conn, project_id)

        feedback_map = defaultdict(dict)

        for f in feedback_rows:
            key = (f["student_name"], f["criterion_name"])
            feedback_map[key]["rationale"] = format_text(f["rationale"])
            feedback_map[key]["keysentence"] = format_text(f["key_sentence"])

        for row in rows:
            student = row["student_name"]

            for key in list(row.keys()):
                if key.endswith("_ai"):
                    criterion = key.rsplit("_", 1)[0]

                    fb = feedback_map.get((student, criterion), {})

                    row[f"{criterion}_rationale"] = fb.get("rationale", "")
                    row[f"{criterion}_keysentence"] = fb.get("keysentence", "")

        rater_groups = defaultdict(list)
        for row in rows:
            rater_groups[row["rater_name"]].append(row)

        output = io.BytesIO()

        with pd.ExcelWriter(output, engine="openpyxl") as writer:

            for rater, data in rater_groups.items():

                multi_rows = []

                for row in data:
                    new_row = {}

                    new_row[("학생", "")] = row["student_name"]

                    for key, value in row.items():

                        if key in ["student_name", "rater_name"]:
                            continue

                        if key.endswith("_human"):
                            c = key.replace("_human", "")
                            new_row[(c, "human")] = value

                        elif key.endswith("_ai"):
                            c = key.replace("_ai", "")
                            new_row[(c, "ai")] = value

                        elif key.endswith("_final"):
                            c = key.replace("_final", "")
                            new_row[(c, "final")] = value

                        elif key.endswith("_rationale"):
                            c = key.replace("_rationale", "")
                            new_row[(c, "rationale")] = value

                        elif key.endswith("_keysentence"):
                            c = key.replace("_keysentence", "")
                            new_row[(c, "keysentence")] = value

                    multi_rows.append(new_row)

                df = pd.DataFrame(multi_rows)

                df.columns = pd.MultiIndex.from_tuples(df.columns)

                columns = list(df.columns)

                header_top = [col[0] for col in columns]
                header_bottom = [col[1] for col in columns]

                df.columns = [f"{col[0]}__{col[1]}" for col in columns]

                sheet_name = f"rater_{rater}"[:31]

                df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=2)

                ws = writer.sheets[sheet_name]

                for col_idx, (top, bottom) in enumerate(zip(header_top, header_bottom), start=1):
                    ws.cell(row=1, column=col_idx, value=top)
                    ws.cell(row=2, column=col_idx, value=bottom)

                merge_map = defaultdict(list)

                for i, top in enumerate(header_top):
                    merge_map[top].append(i + 1)

                for top, cols in merge_map.items():
                    if len(cols) > 1:
                        ws.merge_cells(
                            start_row=1,
                            start_column=cols[0],
                            end_row=1,
                            end_column=cols[-1]
                        )

                for row in ws.iter_rows(min_row=1, max_row=2):
                    for cell in row:
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal="center", vertical="center")

                ws.freeze_panes = "A3"

                for col_idx, col in enumerate(ws.columns, start=1):
                    col_letter = get_column_letter(col_idx)
                    header = ws.cell(row=1, column=col_idx).value

                    width = 15

                    if header and "rationale" in str(header):
                        width = 50
                    elif header and "keysentence" in str(header):
                        width = 60
                    elif header == "학생":
                        width = 20

                    ws.column_dimensions[col_letter].width = width

                    for cell in col:
                        if header and (
                            "rationale" in str(header) or "keysentence" in str(header)
                        ):
                            cell.alignment = Alignment(
                                wrapText=True,
                                vertical="top"
                            )
                        else:
                            cell.alignment = Alignment(
                                horizontal="center",
                                vertical="center"
                            )

        output.seek(0)

        return send_file(
            output,
            download_name=f"{project_name}_export.xlsx",
            as_attachment=True,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

@app.get("/student/<project_name>/<student_id>")
def get_student(project_name, student_id):

    engine = get_engine()

    with engine.connect() as conn:

        project_id = get_project_id(conn, project_name)

        row = conn.execute(
            sqlalchemy.text("""
                SELECT *
                FROM studentDB
                WHERE student_name = :id
                AND project_id = :project_id
            """),
            {
                "id": student_id,
                "project_id": project_id
            }
        ).mappings().fetchone()

        if not row:
            return {"error": "student not found"}, 404

        return jsonify(dict(row))
    
@app.get("/students/<project_name>/<student_range>")
def get_students_by_range(project_name, student_range):

    try:
        start_id, end_id = student_range.split("-")
        start_id = int(start_id)
        end_id = int(end_id)
    except ValueError:
        return {"error": "invalid range format. use start-end"}, 400

    engine = get_engine()

    with engine.connect() as conn:

        project_id = get_project_id(conn, project_name)

        rows = conn.execute(
            sqlalchemy.text("""
                SELECT *
                FROM studentDB
                WHERE project_id = :project_id
                AND student_name BETWEEN :start AND :end
                ORDER BY student_id
            """),
            {
                "project_id": project_id,
                "start": start_id,
                "end": end_id
            }
        ).mappings().fetchall()

        if not rows:
            return {"error": "no students found"}, 404

        return [dict(row) for row in rows]

@app.post("/ai_grade")
def ai_grade():

    data = request.get_json(silent=True)

    if data is None:
        return {"success": False, "message": "Invalid JSON"}, 400

    student_id = data["student_id"]
    rater_uid = data["rater_uid"]
    rater_name = data["rater_name"]
    project_name = data["project_name"]
    student_answer = data["student_answer"]

    criteria_data = data["criteria"]

    engine = get_engine()

    with engine.begin() as conn:

        project_id = get_project_id(conn, project_name)

        project = conn.execute(
            sqlalchemy.text("""
                SELECT prompt_text
                FROM projectDB
                WHERE project_id = :pid
            """),
            {"pid": project_id}
        ).mappings().fetchone()

        prompt_text = project["prompt_text"] if project else ""

        criteria_list = [c["name"] for c in criteria_data]

        # AI 채점 실행
        ai_result = run_ai_grading(student_answer, prompt_text, criteria_list)

        print("AI RESULT RAW:", ai_result)

        ai_scores = ai_result.get("scores", {})
        ai_rationales = ai_result.get("rationales", {})
        ai_keys = ai_result.get("key_sentences", {})

        mapped_scores = {}
        mapped_rationales = {}
        mapped_keys = {}

        ai_keys_list = list(ai_scores.keys())

        for i, criterion in enumerate(criteria_list):

            if i < len(ai_keys_list):
                ai_key = ai_keys_list[i]

                mapped_scores[criterion] = ai_scores.get(ai_key)
                mapped_rationales[criterion] = ai_rationales.get(ai_key, [])
                mapped_keys[criterion] = ai_keys.get(ai_key, [])

            else:
                mapped_scores[criterion] = None
                mapped_rationales[criterion] = []
                mapped_keys[criterion] = []

        ai_result["scores"] = mapped_scores
        ai_result["rationales"] = mapped_rationales
        ai_result["key_sentences"] = mapped_keys

        MODEL_VERSION = "gemini-2.5-flash"

        if not isinstance(ai_scores, dict):
            ai_scores = {}

        # AI 로그 저장
        for criterion, score in mapped_scores.items():

            rationale_list = mapped_rationales.get(criterion, [])
            key_sentence_list = mapped_keys.get(criterion, [])

            rationale_text = "\n".join(rationale_list)
            key_sentence_text = "\n".join(key_sentence_list)

            conn.execute(
                sqlalchemy.text("""
                    INSERT INTO ai_feedback_log
                    (
                        log_id,
                        student_id,
                        criterion_name,
                        score,
                        rationale,
                        key_sentence,
                        model_name,
                        raw_response,
                        created_at,
                        project_id,
                        rater_uid,
                        rater_name
                    )
                    VALUES
                    (
                        :log_id,
                        :student_id,
                        :criterion_name,
                        :score,
                        :rationale,
                        :key_sentence,
                        :model_name,
                        :raw_response,
                        NOW(),
                        :project_id,
                        :rater_uid,
                        :rater_name
                    )
                """),
                {
                    "log_id": str(uuid.uuid4()),
                    "student_id": student_id,
                    "criterion_name": criterion,
                    "score": score,
                    "rationale": rationale_text,
                    "key_sentence": key_sentence_text,
                    "model_name": MODEL_VERSION,
                    "raw_response": json.dumps(ai_result, ensure_ascii=False),
                    "project_id": project_id,
                    "rater_uid": rater_uid,
                    "rater_name": rater_name
                }
            )

    return {
        "success": True,
        "ai_result": ai_result
    }

@app.post("/login")
def login():
    data = request.json
    rater_id = data.get("rater_id")
    password = data.get("password")
    project_name = data.get("project_name")

    COMMON_PASSWORD = os.environ.get("COMMON_PASSWORD", "000000")
    ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "1111")

    # 관리자 로그인 분기
    if password == ADMIN_PASSWORD:
        return {
            "success": True,
            "role": "admin",
            "rater_uid": "admin",
            "rater_name": "admin"
        }

    # 일반 사용자 비밀번호 검사
    if password != COMMON_PASSWORD:
        return {"success": False, "message": "비밀번호 오류"}

    engine = get_engine()
    with engine.connect() as conn:

        # 일반 사용자만 project 검사
        project_id = get_project_id(conn, project_name)

        if not project_id:
            return {"success": False, "message": "project not found"}


        row = conn.execute(
            sqlalchemy.text("""
                SELECT rater_uid, rater_name
                FROM raterDB
                WHERE rater_name = :rid
            """),
            {"rid": rater_id}
        ).mappings().fetchone()

        # project criteria 가져오기
        project_row = conn.execute(
            sqlalchemy.text("""
                SELECT criteria
                FROM projectDB
                WHERE project_id = :pid
            """),
            {"pid": project_id}
        ).mappings().fetchone()

        criteria = project_row["criteria"] if project_row else None

        if row is not None:
            return {
                "success": True,
                "role": "rater",
                "rater_uid": row["rater_uid"],
                "rater_id": row["rater_name"],
                "project_id": project_id,
                "criteria": criteria
            }

        new_uid = conn.execute(
            sqlalchemy.text("SELECT UUID() AS uid")
        ).mappings().fetchone()["uid"]

        conn.execute(
            sqlalchemy.text("""
                INSERT INTO raterDB (rater_uid, rater_name)
                VALUES (:uid, :rid)
            """),
            {"uid": new_uid, "rid": rater_id}
        )
        conn.commit()

        return {
            "success": True,
            "role": "rater",
            "rater_uid": new_uid,
            "rater_id": rater_id,
            "project_id": project_id,
            "criteria": criteria
        }
    
@app.delete("/delete_project/<project_name>")
def delete_project(project_name):

    engine = get_engine()

    with engine.begin() as conn:

        project_id = get_project_id(conn, project_name)

        if not project_id:
            return {"success": False, "message": "project not found"}, 404

        conn.execute(sqlalchemy.text("""
        DELETE FROM scoreDB
        WHERE project_id = :pid
        """), {"pid": project_id})

        conn.execute(sqlalchemy.text("""
        DELETE FROM studentDB
        WHERE project_id = :pid
        """), {"pid": project_id})

        conn.execute(sqlalchemy.text("""
        DELETE FROM projectDB
        WHERE project_id = :pid
        """), {"pid": project_id})

    return {
        "success": True,
        "message": f"Project {project_name} 데이터 삭제 완료"
    }

@app.post("/add_final_score")
def add_final_score():

    data = request.json

    score_id = str(uuid.uuid4())
    student_id = data["student_id"]
    rater_uid = data["rater_uid"]
    rater_name = data["rater_name"]
    project_name = data["project_name"]

    expert_scores = data.get("expert_scores", {})
    ai_scores = data.get("ai_scores", {})
    final_scores = data.get("final_scores", {})

    engine = get_engine()

    with engine.begin() as conn:

        project_id = get_project_id(conn, project_name)

        conn.execute(
            sqlalchemy.text("""
                INSERT INTO scoreDB
                (score_id, student_id, rater_uid, rater_name,
                 stage, scores, project_id, created_at)
                VALUES
                (:score_id, :student_id, :rater_uid, :rater_name,
                 'human', :scores, :project_id, NOW())
            """),
            {
                "score_id": score_id,
                "student_id": student_id,
                "rater_uid": rater_uid,
                "rater_name": rater_name,
                "scores": json.dumps(expert_scores, ensure_ascii=False),
                "project_id": project_id
            }
        )

        conn.execute(
            sqlalchemy.text("""
                INSERT INTO scoreDB
                (score_id, student_id, rater_uid, rater_name,
                 stage, scores, project_id, created_at)
                VALUES
                (:score_id, :student_id, :rater_uid, :rater_name,
                 'ai', :scores, :project_id, NOW())
            """),
            {
                "score_id": score_id,
                "student_id": student_id,
                "rater_uid": rater_uid,
                "rater_name": rater_name,
                "scores": json.dumps(ai_scores, ensure_ascii=False),
                "project_id": project_id
            }
        )

        conn.execute(
            sqlalchemy.text("""
                INSERT INTO scoreDB
                (score_id, student_id, rater_uid, rater_name,
                 stage, scores, project_id, created_at)
                VALUES
                (:score_id, :student_id, :rater_uid, :rater_name,
                 'final', :scores, :project_id, NOW())
            """),
            {
                "score_id": score_id,
                "student_id": student_id,
                "rater_uid": rater_uid,
                "rater_name": rater_name,
                "scores": json.dumps(final_scores, ensure_ascii=False),
                "project_id": project_id
            }
        )

    return {"status": "ok"}


# 프런트엔드 서빙
@app.route("/")
def serve_index():
    return send_from_directory(FRONTEND_BUILD_PATH, "index.html")

@app.route("/<path:path>")
def serve_react(path):

    # API 경로는 React로 보내지 않음
    if path.startswith("api") or path.startswith("ai_grade") or path.startswith("student"):
        return {"error": "API route not found"}, 404

    file_path = os.path.join(FRONTEND_BUILD_PATH, path)

    if os.path.exists(file_path):
        return send_from_directory(FRONTEND_BUILD_PATH, path)

    return send_from_directory(FRONTEND_BUILD_PATH, "index.html")